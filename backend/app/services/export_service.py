from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from app.models.encuesta import Encuesta
from app.models.usuario import Usuario
from app.models.alerta import Alerta
import io

class ExportService:
    
    @staticmethod
    def export_to_excel(db: Session, filtros: dict = None) -> bytes:
        """
        Exporta encuestas a Excel
        
        Args:
            db: Sesión de base de datos
            filtros: Diccionario con filtros opcionales
        
        Returns:
            Bytes del archivo Excel
        """
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Encuestas Bienestar"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "ID Encuesta",
            "Fecha",
            "Tipo Usuario",
            "Documento",
            "Nombres",
            "Apellidos",
            "Programa/Cargo",
            "Promoción",
            "Pregunta 1",
            "Pregunta 2",
            "Pregunta 3",
            "Pregunta 4",
            "Pregunta 5",
            "Puntaje Raw",
            "Puntaje Final",
            "Es Alerta",
            "Estado Alerta",
            "Comentario"
        ]
        
        # Escribir headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Query encuestas
        query = db.query(Encuesta).join(Usuario)
        
        # Aplicar filtros si existen
        if filtros:
            if filtros.get('tipo_usuario'):
                query = query.filter(Usuario.tipo_usuario == filtros['tipo_usuario'])
            if filtros.get('programa'):
                query = query.filter(Usuario.programa == filtros['programa'])
            if filtros.get('es_alerta') is not None:
                query = query.filter(Encuesta.es_alerta == filtros['es_alerta'])
        
        encuestas = query.order_by(Encuesta.created_at.desc()).all()
        
        # Escribir datos
        for row_num, encuesta in enumerate(encuestas, start=2):
            usuario = encuesta.usuario
            
            # Obtener respuestas
            respuestas = {r.pregunta_numero: r.valor for r in encuesta.respuestas}
            
            # Obtener estado de alerta
            alerta = db.query(Alerta).filter(Alerta.encuesta_id == encuesta.id).first()
            estado_alerta = alerta.estado if alerta else "N/A"
            
            # Programa o Cargo según tipo de usuario
            programa_cargo = usuario.programa if usuario.tipo_usuario == "estudiante" else usuario.cargo
            promocion = usuario.promocion if usuario.tipo_usuario == "estudiante" else ""
            
            # Escribir fila
            row_data = [
                encuesta.id,
                encuesta.completed_at.strftime("%d/%m/%Y %H:%M") if encuesta.completed_at else "",
                usuario.tipo_usuario,
                f"{usuario.tipo_documento} {usuario.numero_documento}",
                usuario.nombres,
                usuario.apellidos,
                programa_cargo,
                promocion,
                respuestas.get(1, ""),
                respuestas.get(2, ""),
                respuestas.get(3, ""),
                respuestas.get(4, ""),
                respuestas.get(5, ""),
                encuesta.puntaje_raw,
                encuesta.puntaje_final,
                "SÍ" if encuesta.es_alerta else "NO",
                estado_alerta,
                encuesta.comentario or ""
            ]
            
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
